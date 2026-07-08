import os
import re
import tkinter as tk
from tkinter import filedialog
from typing import Dict, List, Tuple

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    from openpyxl import load_workbook
except ModuleNotFoundError as e:
    raise SystemExit(
        f"缺少依赖包 {e.name}。请先安装依赖：pip install -r requirements.txt"
    ) from e


# 中文字体设置，避免乱码
plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimSun", "Arial Unicode MS"]
plt.rcParams["axes.unicode_minus"] = False


DIMENSION_ORDER = [
    ("B1", "使用对象多元性（B1）", ["使用对象多元性", "使用对象多元性(B1)", "使用对象多元性（B1）", "B1"]),
    ("B2", "管理控制性（B2）", ["管理控制性", "管理控制性(B2)", "管理控制性（B2）", "B2"]),
    ("B3", "公共服务与维护性（B3）", ["公共服务与维护性", "公共服务与维护性(B3)", "公共服务与维护性（B3）", "B3"]),
    ("B4", "物理配置与空间开放性（B4）", ["物理配置与空间开放性", "物理配置与空间开放性(B4)", "物理配置与空间开放性（B4）", "B4"]),
    ("B5", "活动参与与公共活力（B5）", ["活动参与与公共活力", "活动参与与公共活力(B5)", "活动参与与公共活力（B5）", "B5"]),
]

DIMENSION_COLS = [code for code, _, _ in DIMENSION_ORDER]


def parse_score(value):
    """将评分字段转换为数值，跳过空值、N/A、NA、不适用等无效项。"""
    if pd.isna(value):
        return np.nan
    if isinstance(value, str):
        s = value.strip()
        if s == "" or s.lower() in {"n/a", "na", "不适用", "none", "null", "—", "-"}:
            return np.nan
        try:
            return float(s)
        except ValueError:
            return np.nan
    try:
        return float(value)
    except (TypeError, ValueError):
        return np.nan


def read_score_sheet(file_path, sheet_name):
    """读取单张工作表，返回标题、清洗后的数据表和列信息。"""
    raw_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    if raw_df.empty:
        raise ValueError(f"工作表 {sheet_name} 为空。")

    title = raw_df.iloc[0, 0] if raw_df.shape[1] > 0 else sheet_name
    title = title if pd.notna(title) else sheet_name

    if raw_df.shape[0] < 2:
        raise ValueError(f"工作表 {sheet_name} 数据行不足，至少需要 2 行表头。")

    header_row = raw_df.iloc[1].tolist()
    data_df = raw_df.iloc[2:].copy() if raw_df.shape[0] > 2 else pd.DataFrame()

    headers = []
    for i, h in enumerate(header_row):
        if pd.isna(h):
            headers.append(f"col_{i}")
        else:
            headers.append(str(h).strip())

    if data_df.empty:
        data_df = pd.DataFrame(columns=headers)
    else:
        if len(headers) < data_df.shape[1]:
            headers = headers + [f"col_{i}" for i in range(len(headers), data_df.shape[1])]
        elif len(headers) > data_df.shape[1]:
            headers = headers[: data_df.shape[1]]
        data_df.columns = headers

    dim_col = None
    score_col = None

    for col in data_df.columns:
        if dim_col is None and ("维度" in str(col)):
            dim_col = col
        if score_col is None and ("评分" in str(col)):
            score_col = col

    if dim_col is None:
        dim_col = data_df.columns[0]
    if score_col is None:
        score_col = data_df.columns[-1]

    # 将“维度（B）”列向下填充
    data_df[dim_col] = data_df[dim_col].apply(lambda x: np.nan if pd.isna(x) or str(x).strip() == "" else str(x).strip())
    data_df[dim_col] = data_df[dim_col].ffill()

    # 将评分列转换为数值
    data_df[score_col] = data_df[score_col].apply(parse_score)

    return {
        "title": title,
        "sheet_name": sheet_name,
        "df": data_df,
        "dim_col": dim_col,
        "score_col": score_col,
    }


def clean_space_title(raw_title, clean=True):
    """从 Excel 标题中提取空间名称，可选择是否清洗标题。"""
    if not isinstance(raw_title, str):
        raw_title = str(raw_title)
    title = raw_title.strip()
    if not clean:
        return title

    title = title.replace("（", "(").replace("）", ")")
    title = re.sub(r"^(文化事件介入前|文化事件介入后|文化事件介入|介入前|介入后|前|后)\s*[_-]?", "", title)
    title = re.sub(r"(公共性指标评分表|公共性评价表|指标评分表|评分表)$", "", title)
    title = re.sub(r"[\s_\-]+", "", title)
    title = title.replace("公共性", "")
    title = title.strip()
    return title if title else raw_title.strip()


def calculate_dimension_scores(df, space_name=""):
    """根据“维度（B）”和“评分”列计算五个维度的平均分。"""
    if df.empty:
        raise ValueError("输入数据为空，无法计算维度得分。")

    # 自动识别维度列和评分列
    dim_col = None
    score_col = None
    for col in df.columns:
        if dim_col is None and ("维度" in str(col)):
            dim_col = col
        if score_col is None and ("评分" in str(col)):
            score_col = col

    if dim_col is None or score_col is None:
        raise ValueError("无法识别维度列或评分列。")

    def normalize_dim(value):
        if pd.isna(value):
            return ""
        return re.sub(r"\s+", "", str(value).replace("（", "(").replace("）", ")"))

    scores = {}
    for code, label, aliases in DIMENSION_ORDER:
        matched = df[dim_col].apply(lambda x: normalize_dim(x) in {normalize_dim(a) for a in aliases})
        values = df.loc[matched, score_col]
        valid_values = [float(v) for v in values if pd.notna(v) and isinstance(v, (int, float)) and 1 <= float(v) <= 5]
        valid_values += [float(v) for v in values if pd.notna(v) and isinstance(v, str) and re.fullmatch(r"\d+(\.\d+)?", v.strip()) and 1 <= float(v.strip()) <= 5]

        if len(valid_values) == 0:
            print(f"提示：{space_name or '当前空间'} 的维度 {label} 缺失或无有效评分。")
            scores[code] = np.nan
        else:
            scores[code] = float(np.mean(valid_values))

    return scores


def polar_to_xy(radius, angle):
    """将极坐标转换为平面坐标。"""
    return radius * np.cos(angle), radius * np.sin(angle)


def plot_star_model(space_name, scores, save_path):
    """绘制标准五角星结构的公共性评价图，每个维度使用独立四边形色块。"""
    fig, ax = plt.subplots(figsize=(8, 8), dpi=300)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    ax.set_aspect("equal")
    ax.set_xlim(-6.2, 6.2)
    ax.set_ylim(-6.2, 6.2)
    ax.axis("off")

    # 五条维度轴线的角度，第一条位于正上方，顺时针排列
    angles = np.pi / 2 - np.arange(5) * 2 * np.pi / 5
    axis_labels = [
        "使用对象多元性（B1）",
        "管理控制性（B2）",
        "公共服务与维护性（B3）",
        "物理配置与空间开放性（B4）",
        "活动参与与公共活力（B5）",
    ]
    colors = ["#F5A623", "#F36C3D", "#4DB6AC", "#5DADE2", "#A569BD"]

    outer_r = 5.0
    inner_r = 2.3

    # 计算五个凸点和五个凹点
    outer_points = []
    inner_points = []
    for angle in angles:
        x, y = polar_to_xy(outer_r, angle)
        outer_points.append((x, y))

    for i, angle in enumerate(angles):
        next_angle = angles[(i + 1) % len(angles)]
        mid_angle = (angle + next_angle) / 2
        x, y = polar_to_xy(inner_r, mid_angle)
        inner_points.append((x, y))

    # 绘制轴线和 1-5 等级点
    for angle in angles:
        x, y = polar_to_xy(outer_r, angle)
        ax.plot([0, x], [0, y], color="black", lw=1.0, alpha=0.9, zorder=1)
        for r in range(1, 6):
            px, py = polar_to_xy(r, angle)
            ax.scatter(px, py, s=8, color="black", zorder=3)

    # 绘制标准五角星外轮廓：凸点 -> 凹点 -> 凸点 -> ...
    star_xy = []
    for i in range(len(outer_points)):
        star_xy.append(outer_points[i])
        star_xy.append(inner_points[i])
    star_xy.append(outer_points[0])
    star_x = [p[0] for p in star_xy]
    star_y = [p[1] for p in star_xy]
    ax.plot(star_x, star_y, color="black", lw=1.3, zorder=4)

    # 计算每个维度的数值点
    data_points = []
    data_radii = []
    for i, (code, _, _) in enumerate(DIMENSION_ORDER):
        score = scores.get(code, np.nan)
        if pd.notna(score):
            radius = float(score)
            if radius < 0:
                radius = 0
            if radius > 5:
                radius = 5
        else:
            radius = 0.0
        data_radii.append(radius)
        x, y = polar_to_xy(radius, angles[i])
        data_points.append((x, y))

    # 为每个维度绘制独立四边形色块：中心点 -> 左凹点 -> 当前维度数值点 -> 右凹点 -> 回到中心点
    for i in range(5):
        left_idx = (i - 1) % 5
        right_idx = i
        center = (0.0, 0.0)
        left_concave = inner_points[left_idx]
        value_point = data_points[i]
        right_concave = inner_points[right_idx]
        quad_points = [center, left_concave, value_point, right_concave]
        quad_x = [p[0] for p in quad_points]
        quad_y = [p[1] for p in quad_points]
        ax.fill(quad_x, quad_y, color=colors[i], alpha=0.25, zorder=1)

    # 绘制数据点和数值标签
    for i, (x, y) in enumerate(data_points):
        ax.scatter(x, y, s=45, color="black", zorder=5)
        label_x, label_y = x * 1.16, y * 1.16
        ax.text(label_x, label_y, f"{data_radii[i]:.2f}", ha="center", va="center", fontsize=8, fontweight="bold", color="black")

    # 维度名称标签
    for label, angle in zip(axis_labels, angles):
        x, y = polar_to_xy(outer_r + 0.5, angle)
        ax.text(x, y, label, ha="center", va="center", fontsize=8, fontweight="bold", color="black")

    # 左上角标题
    ax.text(
        -5.55,
        5.55,
        space_name,
        ha="left",
        va="top",
        fontsize=12,
        fontweight="bold",
        color="white",
        bbox=dict(boxstyle="square,pad=0.25", facecolor="black", edgecolor="black"),
    )

    plt.tight_layout()
    fig.savefig(save_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def batch_generate_star_charts(file_path, output_folder, clean_title=True):
    """批量读取 Excel 中所有工作表，并生成所有空间的星形图与汇总表。"""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    if not os.path.isabs(file_path):
        file_path = os.path.join(base_dir, file_path)

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

    if not os.path.isabs(output_folder):
        output_folder = os.path.join(base_dir, output_folder)

    os.makedirs(output_folder, exist_ok=True)

    wb = load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    summary_rows = []
    for sheet_name in sheet_names:
        sheet_result = read_score_sheet(file_path, sheet_name)
        raw_title = sheet_result["title"]
        space_name = clean_space_title(raw_title, clean=clean_title)
        df = sheet_result["df"]
        scores = calculate_dimension_scores(df, space_name=space_name)

        safe_name = re.sub(r"[^一-龥A-Za-z0-9._-]+", "_", space_name)
        image_path = os.path.join(output_folder, f"{safe_name}.png")
        plot_star_model(space_name, scores, image_path)
        print(f"已生成: {image_path}")

        row = {"空间名称": space_name}
        for code, label, _ in DIMENSION_ORDER:
            row[label] = scores.get(code, np.nan)
        summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows) 
    # 统一列顺序
    ordered_cols = ["空间名称"]
    for code, label, _ in DIMENSION_ORDER:
        ordered_cols.append(label)
    summary_df = summary_df.reindex(columns=ordered_cols, fill_value=np.nan)
    summary_path = os.path.join(output_folder, "星形图维度得分汇总.xlsx")
    summary_df.to_excel(summary_path, index=False)
    print(f"已生成汇总表: {summary_path}")

    return summary_df


def choose_file_and_run():
    """弹出文件选择窗口，让用户手动选择要读取的 Excel 文件。"""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="请选择 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
    )
    root.destroy()

    if not file_path:
        print("未选择文件，程序已退出。")
        return

    output_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "星形图输出")
    batch_generate_star_charts(file_path=file_path, output_folder=output_folder)


if __name__ == "__main__":
    choose_file_and_run()
